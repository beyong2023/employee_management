from django.shortcuts import render, redirect
from app_1 import models
from django.http import HttpResponse

""" 部门 """


def depart_list(request):
    """部门列表"""
    queryset = models.Department.objects.all()
    return render(request, 'depart_list.html', {'queryset': queryset})


def depart_add(request):
    """添加部门"""
    if request.method == "GET":
        return render(request, 'depart_add.html')
    title = request.POST.get("title")
    models.Department.objects.create(title=title)
    return redirect("/depart/list/")


def depart_delete(request):
    """删除部门"""
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")


def depart_edit(request, nid):
    """修改部门"""
    if request.method == "GET":
        row_object = models.Department.objects.filter(id=nid).first()
        print(row_object.id, row_object.title)
        return render(request, 'depart_edit.html', {"row_object": row_object})
    title = request.POST.get("title")
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list/")



from openpyxl import load_workbook


def depart_multi(request):
    """ 批量上传 """
    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # print(type(file_object))

    # 2.对象传递给openpyxl,由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # cell = sheet.cell(1, 1)
    # print(cell.value)
    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.create(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    return HttpResponse('上传文件')



























